"""
FileName: api.py
Description: 
Time: 2020/9/11 16:42
Project: CiLog
Author: Shurui Gui
"""
import logging
from logging.handlers import SMTPHandler
from cilog.logger import CustomLogger, s2l
from cilog.handler import CustomFileHandler
from cilog.formatter import CustomFormatter, FileFormatter
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os
import copy
import re
import sys
import time





def create_logger(name: str = 'CiLog',
                  file: str = None,
                  enable_mail: bool = False,
                  use_color: bool = True,
                  sub_print: bool = False,
                  **kwargs) -> CustomLogger:
    """
    Create Custom Cilog for logging.

    Args:
        name : str : logger name
        file : str : File path
        enable_mail : bool : Default False
        use_color : bool : Signal for using colored info. Default True
        sub_print : bool : Substitute builtins.print [Advance Usage]
        **kwargs: Dict : Please refer to  the following options (hiden).

        file_mode : str - File open mode. Default: 'a'
        file_level : Literal['DEBUG', 'INFO', 'WARNING', 'IMPORTANT', 'ERROR', 'CRITICAL', 'MAIL'] - Default 'INFO'
        mail_level : Literal['DEBUG', 'INFO', 'WARNING', 'IMPORTANT', 'ERROR', 'CRITICAL', 'MAIL'] - Default 'MAIL'
        mail_setting : dir - Required if enable_mail == True
            {
                mailhost:   string or tuple - YourMailHost or (host, port),
                fromaddr:   string          - YourSenderAddress,
                toaddrs:    list(string)    - List of YourTargetAddresses,
                subject:    string          - Mail Subject,
                credentials:tuple           - (YourUsername, YourPassword),
                secure:     tuple           - () or (KeyfileName) or (KeyfileName, CertificatefileName)
                                                use the secure protocol (TLS),
                timeout:    float           - Default 1.0
            }
        stack_level : Literal['DEBUG', 'INFO', 'WARNING', 'IMPORTANT', 'ERROR', 'CRITICAL', 'MAIL'] - Default 'ERROR'
        msg_fmt : Dict{'DEBUG': debug_fmt, 'INFO': info_fmt, 'WARNING': warning_fmt, 'IMPORTANT': important_fmt,
        'ERROR': error_fmt, 'CRITICAL': critical_fmt, 'MAIL': mail_fmt} - Custom design massage format.
        Please refer to CustomFormatter and url: https://docs.python.org/3/library/logging.html#logrecord-attributes

    Return:
        logger : logging.Logger
    """

    kwargs['file_mode'] = kwargs.get('file_mode') or 'a'
    kwargs['file_level'] = s2l.get(kwargs.get('file_level') or 'INFO')
    kwargs['mail_level'] = s2l.get(kwargs.get('mail_level') or 'MAIL')
    kwargs['mail_setting'] = kwargs.get('mail_setting') or None
    kwargs['stack_level'] = s2l.get(kwargs.get('stack_level') or 'ERROR')
    kwargs['msg_fmt'] = kwargs.get('msg_fmt') or None


    # create logger
    logging.Logger.manager.setLoggerClass(CustomLogger) # set logging.Logger to CustomLogger
    logger: CustomLogger = logging.getLogger(name)
    logger.setLevel(logging.DEBUG)

    # console handler and its formatter
    ch = logging.StreamHandler()
    ch.setLevel(logging.DEBUG)
    cformatter = CustomFormatter(use_color=use_color,
                                 stack_level=kwargs['stack_level'],
                                 msg_fmt=kwargs['msg_fmt'])
    ch.setFormatter(cformatter)
    logger.addHandler(ch)

    # file handler
    if file:
        fh = CustomFileHandler(file, mode=kwargs['file_mode'])
        fh.setLevel(kwargs['file_level'])
        fformatter = FileFormatter(stack_level=kwargs['stack_level'], msg_fmt=kwargs['msg_fmt'])
        fh.setFormatter(fformatter)
        logger.addHandler(fh)

    if enable_mail:
        assert not(kwargs['mail_setting'] is None) and isinstance(kwargs['mail_setting'], dict), "Mail setting is needed."
        setting = kwargs['mail_setting']
        mh = SMTPHandler(setting.get('mailhost'), setting.get('fromaddr'), setting.get('toaddrs'),
                         setting.get('subject'), credentials=setting.get('credentials'),
                         secure=setting.get('secure'), timeout=setting.get('timeout'))
        mh.setLevel(kwargs['mail_level'])
        mformatter = FileFormatter(stack_level=kwargs['stack_level'], msg_fmt=kwargs['msg_fmt'])
        mh.setFormatter(mformatter)
        logger.addHandler(mh)

    if sub_print:
        logger.substitute_print()
        logging.StreamHandler.terminator = ''
        return


    return logger


def json_mail_setting(mail_setting):
    mail_setting["mailhost"] = tuple(mail_setting["mailhost"])
    mail_setting["credentials"] = tuple(mail_setting["credentials"])
    return mail_setting

class timeit(object):

    def __init__(self, time_name):
        self.time_name = time_name

    def __enter__(self):
        self.start = time.time()

    def __exit__(self, exc_type, exc_val, exc_tb):
        print(f'{self.time_name}: {time.time() - self.start}s')


def fill_table(excel_path, value, x, y, z=None, table_format: list=None, *args, **kwargs):
    file = os.path.abspath(os.fspath(excel_path))
    file_path, file_name_ext = os.path.split(file)
    if not os.path.exists(file_path):
        os.makedirs(file_path)
    if not os.path.exists(file):
        excel = Workbook()
        assert 1 <= len(table_format) <= 3, "table_format should have length l should be 1 <= l <= 3, " \
                                       "[row label list, column label list, sheet label list]"
        if len(table_format) < 3:
            table_format += [None] * (3 - len(table_format))

        # --- first sheet ---
        first_sheet = excel.worksheets[0]

        # --- row labels ---
        for ri, row_label in enumerate(table_format[0]):
            if table_format[1] is not None:
                row_idx = ri + 2
            else:
                row_idx = ri + 1
            first_sheet.cell(row_idx, 1).value = row_label

        # --- column labels ---
        if table_format[1] is not None:
            for ci, column_label in enumerate(table_format[1]):
                column_idx = ci + 2
                first_sheet.cell(1, column_idx).value = column_label

            # --- sheet labels ---
            if table_format[2] is not None:
                for si, sheet_label in enumerate(table_format[2]):
                    if si == 0:
                        excel.worksheets[0].title = sheet_label
                        continue
                    excel.copy_worksheet(first_sheet).title = sheet_label

        excel.save(file)
        excel.close()

    # --- add value ---
    excel = load_workbook(file)
    if len(table_format) < 3:
        target_sheet = excel.worksheets[0]
    else:
        try:
            target_sheet = excel.get_sheet_by_name(z)
        except Exception as e:
            raise Exception(f'{e}\n'
                  f'{z} is not in the table_format[2].')

    if table_format[1] is None:
        try:
            row_idx = table_format[0].index(x) + 1
        except Exception as e:
            raise Exception(f'{e}\n'
                            f'{x} is not in the table_format[0].')
        target_sheet.cell(row_idx, 1).value = x
        column_idx = 2
    else:
        try:
            row_idx = table_format[0].index(x) + 2
        except Exception as e:
            raise Exception(f'{e}\n'
                            f'{x} is not in the table_format[0].')
        try:
            column_idx = table_format[1].index(y) + 2
        except Exception as e:
            raise Exception(f'{e}\n'
                            f'{y} is not in the table_format[1].')
        target_sheet.cell(row_idx, 1).value = x
        target_sheet.cell(1, column_idx).value = y

    target_sheet.cell(row_idx, column_idx).value = value
    excel.save(file)
    excel.close()









